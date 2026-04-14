"""
Compute precision, recall, F1 from eval_results_export.csv and DW_merged_results.xlsx.
Produces two output files:
  - eval_run_summary.csv   (per-run metrics for all 18 Ollama runs + ChatGPT + baselines)
  - eval_full_matrix.csv   (per-file predictions across all runs, for cross-checking)
"""

import csv
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
REPO_ROOT = os.path.dirname(PARENT_DIR)

CSV_PATH = os.path.join(PARENT_DIR, "eval_results_export.csv")
XLSX_PATH = os.path.join(REPO_ROOT, "DW_merged_results.xlsx")
OUT_SUMMARY = os.path.join(PARENT_DIR, "eval_run_summary.csv")
OUT_MATRIX = os.path.join(PARENT_DIR, "eval_full_matrix.csv")


def normalize_prediction(val):
    """Normalize Yes/No/DW/yes/no/1/0 to 'Yes' or 'No'."""
    v = str(val).strip().lower()
    if v in ("yes", "dw", "1", "true"):
        return "Yes"
    return "No"


def compute_metrics(ground_truths, predictions):
    """Compute TP, TN, FP, FN, precision, recall, F1, accuracy."""
    tp = fn = fp = tn = 0
    for gt, pred in zip(ground_truths, predictions):
        g = gt == "Yes"
        p = pred == "Yes"
        if g and p:
            tp += 1
        elif g and not p:
            fn += 1
        elif not g and p:
            fp += 1
        else:
            tn += 1
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
    accuracy = (tp + tn) / (tp + tn + fp + fn) if (tp + tn + fp + fn) > 0 else 0.0
    return {
        "TP": tp, "TN": tn, "FP": fp, "FN": fn,
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(f1, 4),
        "accuracy": round(accuracy, 4),
        "total": tp + tn + fp + fn,
    }


def load_eval_csv(path):
    """Load eval_results_export.csv, return dict of run_id -> list of rows."""
    runs = defaultdict(list)
    run_meta = {}
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rid = int(row["run_id"])
            runs[rid].append(row)
            if rid not in run_meta:
                run_meta[rid] = {
                    "model_name": row["model_name"],
                    "created_at": row["created_at"],
                }
    return runs, run_meta


def load_xlsx_chatgpt(path):
    """Load DW_merged_results.xlsx ChatGPT column + baselines."""
    if openpyxl is None:
        print("WARNING: openpyxl not installed, skipping XLSX data")
        return None, None, None
    wb = openpyxl.load_workbook(path)

    # Main sheet: per-file results
    ws = wb["DW_merged_results"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        filename, dynamics, adoctor, manual, chatgpt = row
        rows.append({
            "class_name": str(filename),
            "ground_truth": normalize_prediction(manual),
            "dynamics": normalize_prediction(dynamics),
            "adoctor": normalize_prediction(adoctor),
            "chatgpt": normalize_prediction(chatgpt),
        })

    # Evaluation sheet: pre-computed baselines
    ws2 = wb["Evaluation"]
    baselines = {}
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
        tool = str(row[0])
        baselines[tool] = {
            "TP": int(row[1]), "TN": int(row[2]),
            "FP": int(row[3]), "FN": int(row[4]),
            "precision": round(float(row[5]), 4),
            "recall": round(float(row[6]), 4),
            "f1": round(float(row[7]), 4),
            "accuracy": round(float(row[8]), 4),
        }

    return rows, baselines, wb


def main():
    print(f"Loading eval CSV: {CSV_PATH}")
    runs, run_meta = load_eval_csv(CSV_PATH)

    print(f"Loading XLSX: {XLSX_PATH}")
    xlsx_rows, xlsx_baselines, _ = load_xlsx_chatgpt(XLSX_PATH)

    # ── Compute metrics for each Ollama eval run ──
    summary_rows = []
    for rid in sorted(runs.keys()):
        rows = runs[rid]
        gts = [r["ground_truth"] for r in rows]
        preds = [r["prediction"] for r in rows]
        m = compute_metrics(gts, preds)
        meta = run_meta[rid]
        summary_rows.append({
            "run_id": rid,
            "model_name": meta["model_name"],
            "pattern": "DW",
            "created_at": meta["created_at"],
            **m,
        })

    # ── Compute ChatGPT metrics from XLSX ──
    if xlsx_rows:
        gts = [r["ground_truth"] for r in xlsx_rows]
        preds = [r["chatgpt"] for r in xlsx_rows]
        m = compute_metrics(gts, preds)
        summary_rows.append({
            "run_id": "chatgpt",
            "model_name": "ChatGPT (from XLSX)",
            "pattern": "DW",
            "created_at": "N/A",
            **m,
        })

    # ── Add DynAMICS and aDoctor baselines from XLSX ──
    if xlsx_rows:
        for tool_key, display_name in [("dynamics", "DynAMICS (baseline)"), ("adoctor", "aDoctor (baseline)")]:
            gts = [r["ground_truth"] for r in xlsx_rows]
            preds = [r[tool_key] for r in xlsx_rows]
            m = compute_metrics(gts, preds)
            summary_rows.append({
                "run_id": tool_key,
                "model_name": display_name,
                "pattern": "DW",
                "created_at": "N/A",
                **m,
            })

    # ── Write summary CSV ──
    fields = ["run_id", "model_name", "pattern", "TP", "TN", "FP", "FN",
              "precision", "recall", "f1", "accuracy", "total", "created_at"]
    with open(OUT_SUMMARY, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(summary_rows)
    print(f"\nSaved run summary: {OUT_SUMMARY}")

    # ── Print summary table ──
    print(f"\n{'='*100}")
    print(f"{'Run':>6} | {'Model':<30} | {'TP':>3} {'TN':>3} {'FP':>3} {'FN':>3} | "
          f"{'Prec':>6} {'Recall':>6} {'F1':>6} {'Acc':>6} | {'N':>3}")
    print(f"{'-'*100}")
    for r in summary_rows:
        print(f"{str(r['run_id']):>6} | {r['model_name']:<30} | "
              f"{r['TP']:>3} {r['TN']:>3} {r['FP']:>3} {r['FN']:>3} | "
              f"{r['precision']:>6.4f} {r['recall']:>6.4f} {r['f1']:>6.4f} {r['accuracy']:>6.4f} | "
              f"{r['total']:>3}")
    print(f"{'='*100}")

    # ── Write full per-file matrix ──
    all_files = sorted(set(r["class_name"] for rows in runs.values() for r in rows))
    run_ids_sorted = sorted(runs.keys())

    matrix_fields = ["class_name", "ground_truth"]
    for rid in run_ids_sorted:
        meta = run_meta[rid]
        matrix_fields.append(f"run{rid}_{meta['model_name']}")
    if xlsx_rows:
        matrix_fields.extend(["ChatGPT", "DynAMICS", "aDoctor"])

    # Build lookup
    file_data = {}
    for rid in run_ids_sorted:
        for r in runs[rid]:
            key = r["class_name"]
            if key not in file_data:
                file_data[key] = {"class_name": key, "ground_truth": r["ground_truth"]}
            col = f"run{rid}_{run_meta[rid]['model_name']}"
            file_data[key][col] = r["prediction"]

    if xlsx_rows:
        for r in xlsx_rows:
            key = r["class_name"]
            if key in file_data:
                file_data[key]["ChatGPT"] = r["chatgpt"]
                file_data[key]["DynAMICS"] = r["dynamics"]
                file_data[key]["aDoctor"] = r["adoctor"]

    with open(OUT_MATRIX, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=matrix_fields, extrasaction="ignore")
        w.writeheader()
        for fn in all_files:
            if fn in file_data:
                w.writerow(file_data[fn])
    print(f"Saved full matrix: {OUT_MATRIX}")


if __name__ == "__main__":
    main()
